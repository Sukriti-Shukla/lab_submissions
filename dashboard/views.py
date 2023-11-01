from django.shortcuts import render, redirect
from .forms import ExperimentDataForm
from .models import ExperimentData, Deadline
from datetime import date
from .forms import ExperimentSearchForm
from .forms import DeadlineForm
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from django.urls import reverse
from datetime import datetime
from django.utils import timezone
from .forms import TemplateForm
import pandas as pd
import os
from django.conf import settings
import matplotlib.pyplot as plt
from django.http import JsonResponse
import json
import numpy as np
from django.utils.html import escapejs
from django.utils.html import mark_safe
from django.http import JsonResponse
from openpyxl import load_workbook
from django.shortcuts import render
from plotly.offline import plot
import plotly.graph_objects as go
import dashboard.dash_app
import dashboard.dash_app2
# Create your views here.
def home(request):
    return render(request, 'home.html')

def upload(request):
    if request.method == 'POST':
        form = ExperimentDataForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('home')  # Replace with the name of your home view
    else:
        form = ExperimentDataForm()
    return render(request, 'upload.html', {'form': form})

def download(request):
    form = ExperimentSearchForm(request.GET or None)
    experiments = ExperimentData.objects.all()

    if form.is_valid():
        experiment_number = form.cleaned_data.get("experiment_number")
        matrikel_number = form.cleaned_data.get("matrikel_number")
        semester = form.cleaned_data.get("semester")
        
        if experiment_number:
            experiments = experiments.filter(experiment_number=experiment_number)
        if matrikel_number is not None:
            experiments = experiments.filter(matrikel_number=matrikel_number)
        if semester:
            experiments = experiments.filter(semester=semester)

    context = {
        "form": form,
        "experiments": experiments
    }
    return render(request, "download.html", context)
def analyze(request):
    return render(request, 'analyze.html')

def input_experiment(request):
    if request.method == 'POST':
        form = DeadlineForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')  
    else:
        form = DeadlineForm()
    return render(request, 'input_experiment.html', {'form': form})

def upload_template(request):
    if request.method == "POST":
        form = TemplateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('home')  
    else:
        form = TemplateForm()
    return render(request, 'upload_template.html', {'form': form})



@login_required

def my_submissions(request):
    experiments = ExperimentData.objects.filter(matrikel_number=request.user.username)
    today = datetime.now()
    today = timezone.now()

    for experiment in experiments:
        # Try to get the deadline for the experiment, matrikel_number, and semester
        try:
            deadline_obj = Deadline.objects.get(
                experiment_name=experiment.experiment_number,
                matrikel_number=request.user.username,
                semester=experiment.semester 
            )
            deadline = deadline_obj.experiment_deadline
        except Deadline.DoesNotExist:
            deadline = None

        if deadline:
            experiment.show_update_delete = today <= deadline
        else:
            experiment.show_update_delete = True
    
    context = {
        'experiments': experiments,
    }
    return render(request, 'my_submissions.html', context)

@login_required
def update_experiment(request, experiment_id):
    experiment = ExperimentData.objects.get(id=experiment_id)
    if request.method == 'POST':
        form = ExperimentDataForm(request.POST, request.FILES, instance=experiment)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('my_submissions'))
    else:
        form = ExperimentDataForm(instance=experiment)
    return render(request, 'update_experiment.html', {'form': form})


@login_required
def delete_experiment(request, experiment_id):
    experiment = ExperimentData.objects.get(id=experiment_id)
    experiment.delete()
    return HttpResponseRedirect(reverse('my_submissions'))


def analysis_test(request):
    form = ExperimentSearchForm(request.GET or None)
    experiments = None
    dash_context = {}
    if form.is_valid():
        matrikel_number = form.cleaned_data.get('matrikel_number')
        experiment_number = form.cleaned_data.get('experiment_number')
        semester = form.cleaned_data.get('semester')

        experiments = ExperimentData.objects.all()
        
        if matrikel_number:
            experiments = experiments.filter(matrikel_number=matrikel_number)
        if experiment_number:
            experiments = experiments.filter(experiment_number=experiment_number)
        if semester:
            experiments = experiments.filter(semester=semester)
    context = {'form': form, 'experiments': experiments, 'dash_context': dash_context}
    return render(request, 'analysis_test.html', context)

def analyze_view(request, experiment_id):
    experiment = ExperimentData.objects.get(id=experiment_id)
    raw_data_path = experiment.raw_data.path
    dash_context = {'file-path-output': {'value': raw_data_path}}
    dash_context_json = json.dumps(dash_context)
    print("Dash Context:", dash_context)  # Debugging line
    form = ExperimentSearchForm()  # Initialize an empty form
    experiments = ExperimentData.objects.all()  # or filter as needed
    context = {'form': form, 'experiments': experiments, 'dash_context': dash_context_json}
    print(dash_context_json)
    return render(request, 'test_dash.html', context)

def get_experiment_data(request, experiment_id):
    experiment = ExperimentData.objects.get(id=experiment_id)
    file_path = experiment.raw_data.path
    
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    return JsonResponse({'status': 'success', 'data': data})

def analyze_experiment(request, experiment_id):
    experiment = ExperimentData.objects.get(id=experiment_id)
    file_path = experiment.raw_data.path
    
    if not (file_path.endswith('.xlsx') or file_path.endswith('.xls')):
        return JsonResponse({'status': 'error', 'message': 'Not an Excel file'})
    
    wb = load_workbook(filename=file_path, read_only=True)
    
    if len(wb.sheetnames) != 1:
        return JsonResponse({'status': 'error', 'message': 'More than one sheet present'})
    
    sheet = wb.active
    
    if sheet.max_column < 2:
        return JsonResponse({'status': 'error', 'message': 'Less than 2 columns present'})
    
    # If everything is fine, get column names (assuming they are in the first row)
    column_names = [cell.value for cell in sheet[1]]
    
    return JsonResponse({'status': 'success', 'columns': column_names})


def replace_nan_with_null(d):
    for key, value in d.items():
        if isinstance(value, dict):
            replace_nan_with_null(value)
        elif value != value:  
            d[key] = None


def analyze(request):
    form = ExperimentSearchForm()
    experiments = None
    analysis_data = None
    plots = {}
    sheet_titles = []
    if request.method == 'GET':
        form = ExperimentSearchForm(request.GET)
        if form.is_valid():
            experiment_number = form.cleaned_data['experiment_number']
            matrikel_number = form.cleaned_data['matrikel_number']
            semester = form.cleaned_data['semester']

            experiments = ExperimentData.objects.all()

            if experiment_number:
                experiments = experiments.filter(experiment_number=experiment_number)
            if matrikel_number:
                experiments = experiments.filter(matrikel_number=matrikel_number)
            if semester:
                experiments = experiments.filter(semester=semester)

            if 'analyze_id' in request.GET:
                analyze_id = request.GET['analyze_id']
                experiment_to_analyze = ExperimentData.objects.get(id=analyze_id)
                file_path = experiment_to_analyze.raw_data.path
                
                plot_dir = os.path.join(settings.MEDIA_ROOT, 'plots')
                
                if not os.path.exists(plot_dir):
                    os.makedirs(plot_dir)

                if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                    xls = pd.ExcelFile(file_path)
                    analysis_data = {}
                    sheet_titles = []
                    for sheet_name in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name)
                        sheet_title = df.columns[0]
                        #analysis_data[sheet_name] = df.to_dict()
                        analysis_data[sheet_name] = {'data': df.to_dict(), 'sheet_title': sheet_title}
                        sheet_titles.append(sheet_title)
                        # Check if the sheet has exactly 2 columns
                        if len(df.columns) == 2:
                            # Convert data to float and plot
                            plt.figure()
                            second_row_data = df.iloc[0, :] 
                            plt.plot(df.iloc[2:, 0].astype(float), df.iloc[2:, 1].astype(float))
                            # plt.xlabel(df.columns[0])
                            # plt.ylabel(df.columns[1])
                            plt.xlabel(str(second_row_data[0]))
                            plt.ylabel(str(second_row_data[1]))
                            plt.title(sheet_name)
                            img_path = os.path.join(plot_dir, f"{sheet_name}.png")
                            plt.savefig(img_path)
                            # Store the relative path
                            relative_img_path = os.path.join('plots', f"{sheet_name}.png")
                            plots[sheet_name] = relative_img_path


    return render(request, 'analyze.html', {
        'form': form,
        'experiments': experiments,
        'analysis_data': analysis_data,
        'plots': plots,
        'sheet_titles': sheet_titles
    })

def test_dash(request):
    filepath = 'red'
    dash_context = {'dropdown-color': {'value': filepath}}
    context = {
        'my_var': 'Hello, World!',
        'dash_context': json.dumps(dash_context)

    }
    # filepath = 'red'  # Replace with the actual file path
    # dash_context = {'dropdown-color': {'value': filepath}}
    # context = {'dash_context': json.dumps(dash_context)}
    return render(request, 'test_dash.html', context)
